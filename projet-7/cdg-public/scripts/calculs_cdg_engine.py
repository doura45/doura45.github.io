#!/usr/bin/env python3
"""
Moteur de calculs de pilotage budgétaire retail v2 (Contrôle de Gestion).
Calcule les décompositions d'écarts, l'effet mix, le P&L analytique, le seuil de rentabilité,
les métriques BFR/Cash et les scénarios d'atterrissage.
Génère le classeur Excel complet : Pilotage_Budgetaire_Retail_v2.xlsx
"""

import os
import csv
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def run_cdg_engine():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_path = os.path.join(base_dir, "data", "donnees_budget_realise_retail.csv")
    excel_dir = os.path.join(base_dir, "excel")
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = os.path.join(excel_dir, "Pilotage_Budgetaire_Retail_v2.xlsx")

    # Charger les données CSV
    df = pd.read_csv(data_path)
    
    # Calculs de colonnes dérivées brutes
    df["ca_budget"] = df["qte_budget"] * df["pu_vente_budget"]
    df["ca_reel"] = df["qte_reelle"] * df["pu_vente_reel"]
    df["ecart_ca_total"] = df["ca_reel"] - df["ca_budget"]
    
    df["cogs_budget"] = df["qte_budget"] * df["cout_achat_unit_budget"]
    df["cogs_reel"] = df["qte_reelle"] * df["cout_achat_unit_reel"]
    
    df["marge_brute_budget"] = df["ca_budget"] - df["cogs_budget"]
    df["marge_brute_reel"] = df["ca_reel"] - df["cogs_reel"]
    df["ecart_marge_total"] = df["marge_brute_reel"] - df["marge_brute_budget"]

    # --- 1. DECOMPOSITION DES ECARTS (ANNUEL PAR CATEGORIE) ---
    cat_summary = df.groupby("categorie").agg({
        "qte_budget": "sum",
        "qte_reelle": "sum",
        "ca_budget": "sum",
        "ca_reel": "sum",
        "cogs_budget": "sum",
        "cogs_reel": "sum",
        "marge_brute_budget": "sum",
        "marge_brute_reel": "sum",
        "couts_variables_autres": "sum",
        "couts_fixes_affectes": "sum",
        "stock_debut": "first",
        "stock_fin": "last",
        "creances_clients": "mean",
        "dettes_fournisseurs": "mean"
    }).reset_index()

    # Prix unitaires moyens pondérés et coûts unitaires moyens pondérés
    cat_summary["pu_b"] = cat_summary["ca_budget"] / cat_summary["qte_budget"]
    cat_summary["pu_r"] = cat_summary["ca_reel"] / cat_summary["qte_reelle"]
    cat_summary["cu_b"] = cat_summary["cogs_budget"] / cat_summary["qte_budget"]
    cat_summary["cu_r"] = cat_summary["cogs_reel"] / cat_summary["qte_reelle"]
    
    # Décomposition Écart CA
    cat_summary["ecart_ca_total"] = cat_summary["ca_reel"] - cat_summary["ca_budget"]
    cat_summary["ecart_ca_vol"] = (cat_summary["qte_reelle"] - cat_summary["qte_budget"]) * cat_summary["pu_b"]
    cat_summary["ecart_ca_prix"] = (cat_summary["pu_r"] - cat_summary["pu_b"]) * cat_summary["qte_reelle"]

    # Décomposition Écart Marge
    cat_summary["ecart_marge_total"] = cat_summary["marge_brute_reel"] - cat_summary["marge_brute_budget"]
    cat_summary["ecart_marge_qte"] = (cat_summary["qte_reelle"] - cat_summary["qte_budget"]) * (cat_summary["pu_b"] - cat_summary["cu_b"])
    cat_summary["ecart_marge_prix_vente"] = (cat_summary["pu_r"] - cat_summary["pu_b"]) * cat_summary["qte_reelle"]
    cat_summary["ecart_marge_cost_achat"] = - (cat_summary["cu_r"] - cat_summary["cu_b"]) * cat_summary["qte_reelle"]

    # --- 2. EFFET MIX MULTI-CATEGORIES ---
    Q_tot_b = cat_summary["qte_budget"].sum()
    Q_tot_r = cat_summary["qte_reelle"].sum()
    Marge_tot_b = cat_summary["marge_brute_budget"].sum()
    m_bar_b = Marge_tot_b / Q_tot_b # Marge moyenne pondérée budgétée

    cat_summary["m_b_i"] = cat_summary["pu_b"] - cat_summary["cu_b"]
    cat_summary["mix_b_i"] = cat_summary["qte_budget"] / Q_tot_b
    cat_summary["mix_r_i"] = cat_summary["qte_reelle"] / Q_tot_r
    
    cat_summary["effet_mix_i"] = Q_tot_r * (cat_summary["mix_r_i"] - cat_summary["mix_b_i"]) * (cat_summary["m_b_i"] - m_bar_b)
    
    effet_volume_pur_total = (Q_tot_r - Q_tot_b) * m_bar_b
    effet_mix_total = cat_summary["effet_mix_i"].sum()

    # --- 3. P&L ANALYTIQUE PAR CATEGORIE ---
    cat_summary["mcv"] = cat_summary["marge_brute_reel"] - cat_summary["couts_variables_autres"]
    cat_summary["taux_mcv"] = cat_summary["mcv"] / cat_summary["ca_reel"]
    cat_summary["resultat_analytique"] = cat_summary["mcv"] - cat_summary["couts_fixes_affectes"]
    cat_summary["seuil_rentabilite"] = cat_summary["couts_fixes_affectes"] / cat_summary["taux_mcv"]
    cat_summary["marge_securite"] = cat_summary["ca_reel"] - cat_summary["seuil_rentabilite"]
    cat_summary["indice_securite"] = cat_summary["marge_securite"] / cat_summary["ca_reel"]
    cat_summary["levier_operationnel"] = cat_summary["mcv"] / cat_summary["resultat_analytique"]

    # --- 4. CREATION DU CLASSEUR EXCEL FORMATTE ---
    wb = openpyxl.Workbook()
    
    # Styles prédéfinis
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Navy Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    bold_font = Font(name="Calibri", size=11, bold=True)
    num_fmt_currency = "#,##0 €"
    num_fmt_pct = "0.0%"
    num_fmt_qty = "#,##0"

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # -------------------------------------------------------------
    # ONGLET 1 : Données Brutes
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "1_Donnees_Brutes"
    ws1.append(["Periode", "Categorie", "Qte Budget", "Qte Reelle", "PU Vente Budget", "PU Vente Reel", "CU Achat Budget", "CU Achat Reel", "Couts Var Autres", "Couts Fixes", "Stock Debut", "Stock Fin", "Creances Clients", "Dettes Fournisseurs"])
    
    for row in df.itertuples():
        ws1.append([
            row.periode, row.categorie, row.qte_budget, row.qte_reelle,
            row.pu_vente_budget, row.pu_vente_reel, row.cout_achat_unit_budget, row.cout_achat_unit_reel,
            row.couts_variables_autres, row.couts_fixes_affectes, row.stock_debut, row.stock_fin,
            row.creances_clients, row.dettes_fournisseurs
        ])
    
    # Formater en-tête Onglet 1
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # -------------------------------------------------------------
    # ONGLET 2 : Decomposition Ecarts & Effet Mix
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="2_Decomposition_Ecarts")
    ws2.append(["DECOMPOSITION DES ECARTS DE CA ET DE MARGE (ANNUEL)"])
    ws2[f"A1"].font = title_font
    ws2.append([])

    headers_ecarts = [
        "Catégorie", "CA Budget", "CA Réel", "Écart CA Total", "Écart Volume", "Écart Prix",
        "Marge Budget", "Marge Réelle", "Écart Marge Total", "Écart Quantité", "Écart Prix Vente", "Écart Coût Achat", "Effet Mix"
    ]
    ws2.append(headers_ecarts)
    for cell in ws2[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in cat_summary.itertuples():
        ws2.append([
            row.categorie,
            row.ca_budget,
            row.ca_reel,
            row.ecart_ca_total,
            row.ecart_ca_vol,
            row.ecart_ca_prix,
            row.marge_brute_budget,
            row.marge_brute_reel,
            row.ecart_marge_total,
            row.ecart_marge_qte,
            row.ecart_marge_prix_vente,
            row.ecart_marge_cost_achat,
            row.effet_mix_i
        ])

    # Ligne Total Consolidé
    ws2.append([
        "TOTAL CONSOLIDÉ",
        cat_summary["ca_budget"].sum(),
        cat_summary["ca_reel"].sum(),
        cat_summary["ecart_ca_total"].sum(),
        cat_summary["ecart_ca_vol"].sum(),
        cat_summary["ecart_ca_prix"].sum(),
        cat_summary["marge_brute_budget"].sum(),
        cat_summary["marge_brute_reel"].sum(),
        cat_summary["ecart_marge_total"].sum(),
        cat_summary["ecart_marge_qte"].sum(),
        cat_summary["ecart_marge_prix_vente"].sum(),
        cat_summary["ecart_marge_cost_achat"].sum(),
        effet_mix_total
    ])
    
    last_row_2 = ws2.max_row
    for col in range(1, len(headers_ecarts) + 1):
        cell = ws2.cell(row=last_row_2, column=col)
        cell.font = bold_font
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # -------------------------------------------------------------
    # ONGLET 3 : Compte de Résultat Analytique & Seuil de Rentabilité
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="3_PnL_Analytique")
    ws3.append(["COMPTE DE RÉSULTAT ANALYTIQUE PAR CATÉGORIE"])
    ws3[f"A1"].font = title_font
    ws3.append([])

    headers_pnl = [
        "Catégorie", "Chiffre d'Affaires", "COGS (Achats)", "Marge Brute", "Autres Coûts Var.",
        "Marge / Coûts Var (MCV)", "Taux de MCV", "Coûts Fixes Affectés", "Résultat Analytique",
        "Seuil Rentabilité (CA)", "Point Mort (Jours)", "Levier Opérationnel"
    ]
    ws3.append(headers_pnl)
    for cell in ws3[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in cat_summary.itertuples():
        pm_jours = round((row.seuil_rentabilite / row.ca_reel) * 365)
        ws3.append([
            row.categorie,
            row.ca_reel,
            row.cogs_reel,
            row.marge_brute_reel,
            row.couts_variables_autres,
            row.mcv,
            row.taux_mcv,
            row.couts_fixes_affectes,
            row.resultat_analytique,
            row.seuil_rentabilite,
            pm_jours,
            row.levier_operationnel
        ])

    # Total Consolidé PnL
    ca_tot = cat_summary["ca_reel"].sum()
    cogs_tot = cat_summary["cogs_reel"].sum()
    mb_tot = cat_summary["marge_brute_reel"].sum()
    cva_tot = cat_summary["couts_variables_autres"].sum()
    mcv_tot = cat_summary["mcv"].sum()
    taux_mcv_tot = mcv_tot / ca_tot
    cf_tot = cat_summary["couts_fixes_affectes"].sum()
    res_tot = mcv_tot - cf_tot
    sr_tot = cf_tot / taux_mcv_tot
    pm_tot_jours = round((sr_tot / ca_tot) * 365)
    levier_tot = mcv_tot / res_tot

    ws3.append([
        "TOTAL CONSOLIDÉ", ca_tot, cogs_tot, mb_tot, cva_tot, mcv_tot,
        taux_mcv_tot, cf_tot, res_tot, sr_tot, pm_tot_jours, levier_tot
    ])

    last_row_3 = ws3.max_row
    for col in range(1, len(headers_pnl) + 1):
        cell = ws3.cell(row=last_row_3, column=col)
        cell.font = bold_font
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # -------------------------------------------------------------
    # ONGLET 4 : Stocks, BFR & Cash
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="4_Stock_BFR_Cash")
    ws4.append(["INDICATEURS DE STOCK, BFR ET CASH-FLOW"])
    ws4[f"A1"].font = title_font
    ws4.append([])

    headers_bfr = [
        "Catégorie", "Stock Moyen", "COGS Annuel", "Rotation Stock", "Durée Stock (J)",
        "Coût Possession (20%)", "Créances Clients", "Dettes Fournisseurs", "BFR Net", "DSO (J)", "DPO (J)"
    ]
    ws4.append(headers_bfr)
    for cell in ws4[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in cat_summary.itertuples():
        stock_moyen = (row.stock_debut + row.stock_fin) / 2
        rot_stock = row.cogs_reel / stock_moyen
        duree_stock = round(365 / rot_stock)
        cout_possession = stock_moyen * 0.20
        bfr = stock_moyen + row.creances_clients - row.dettes_fournisseurs
        dso = round((row.creances_clients / (row.ca_reel * 1.20)) * 365)
        dpo = round((row.dettes_fournisseurs / (row.cogs_reel * 1.20)) * 365)

        ws4.append([
            row.categorie, stock_moyen, row.cogs_reel, round(rot_stock, 2), duree_stock,
            cout_possession, row.creances_clients, row.dettes_fournisseurs, bfr, dso, dpo
        ])

    # -------------------------------------------------------------
    # ONGLET 5 : Atterrissage & Rolling Forecast
    # -------------------------------------------------------------
    ws5 = wb.create_sheet(title="5_Atterrissage_Forecast")
    ws5.append(["SCÉNARIOS D'ATTERRISSAGE ANNUEL (ROLLING FORECAST)"])
    ws5[f"A1"].font = title_font
    ws5.append([])

    ws5.append(["Budget Annuel Initial", df["ca_budget"].sum()])
    
    # Cumul 8 mois réel (Janv - Août)
    df_8m = df[df["periode"] <= "2025-08"]
    df_4m_b = df[df["periode"] > "2025-08"]
    
    ca_realise_8m = df_8m["ca_reel"].sum()
    ca_budget_8m = df_8m["ca_budget"].sum()
    ca_budget_4m_restant = df_4m_b["ca_budget"].sum()

    taux_realisation_8m = ca_realise_8m / ca_budget_8m
    
    # Scénario A : Taux de réalisation constaté à date maintenu
    atterrissage_scen_a = ca_realise_8m + (ca_budget_4m_restant * taux_realisation_8m)
    ecart_atterrissage_a = atterrissage_scen_a - df["ca_budget"].sum()

    # Scénario B : Retour à la norme (Budget 4 mois restants inchangé)
    atterrissage_scen_b = ca_realise_8m + ca_budget_4m_restant
    ecart_atterrissage_b = atterrissage_scen_b - df["ca_budget"].sum()

    ws5.append(["Réalisé Cumulé à date (8 mois)", ca_realise_8m])
    ws5.append(["Budget Cumulé à date (8 mois)", ca_budget_8m])
    ws5.append(["Taux de réalisation à date", taux_realisation_8m])
    ws5.append([])

    ws5.append(["Scénario", "Projection Annuelle", "Écart vs Budget", "Hypothèse"])
    ws5.append(["Scénario A (Tendanciel)", atterrissage_scen_a, ecart_atterrissage_a, "Poursuite du taux de réalisation constaté à date sur le T4"])
    ws5.append(["Scénario B (Retour Norme)", atterrissage_scen_b, ecart_atterrissage_b, "Réalisation à 100% du budget initial sur le T4"])

    # Ajustement largeur automatique pour toutes les feuilles
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Sauvegarder le classeur Excel
    wb.save(excel_path)
    print(f"✅ Classeur Excel structuré généré : {excel_path}")
    return cat_summary, effet_volume_pur_total, effet_mix_total

if __name__ == "__main__":
    run_cdg_engine()
